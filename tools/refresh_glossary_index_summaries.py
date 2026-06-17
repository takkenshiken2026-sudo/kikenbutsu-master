#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
glossary_terms.csv の index_summary（用語一覧・概要）を下書き生成・レビュー出力・反映する。

  python3 tools/refresh_glossary_index_summaries.py --export-review ~/Desktop/用語一覧_index_summary_レビュー.xlsx
  python3 tools/refresh_glossary_index_summaries.py --draft --dry-run
  python3 tools/refresh_glossary_index_summaries.py --draft --apply
  python3 tools/refresh_glossary_index_summaries.py --apply-from ~/Desktop/用語一覧_index_summary_レビュー.xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.build_glossary_pages import load_glossary_entries, terms_index_snippet  # noqa: E402
from tools.glossary_index_summary_rules import (  # noqa: E402
    INDEX_SUMMARY_MAX_LEN,
    INDEX_SUMMARY_MIN_LEN,
    audit_index_summary_cross_rows,
    check_index_summary_row,
    cross_conflicts_with,
    index_summary_fill_stats,
    normalize_key,
    norm,
)
from tools.index_summary_utils import (  # noqa: E402
    _accept_candidate,
    _break_cross_similarity,
    _build_substance_summary,
    _minimal_exam_summary,
    _push_below_cross_threshold,
    _register_index_summary,
    _dedupe_draft,
    draft_glossary_index_summary,
    resolve_unique_index_summary,
)

CSV_PATH = ROOT / "data" / "glossary_terms.csv"

_CROSS_PARTNER_RE = re.compile(r"index_summary が '([^']+)' と高類似")

REVIEW_COLUMNS = (
    "term",
    "category",
    "importance",
    "index_summary_current",
    "index_summary_draft",
    "index_summary_final",
    "legacy_snippet",
    "legal_basis",
    "exam_points",
    "faq_1_excerpt",
    "draft_len",
    "draft_status",
    "review_status",
)


def load_rows() -> tuple[list[str], list[dict[str, str]]]:
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        return fieldnames, list(reader)


def save_rows(fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    with CSV_PATH.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n", extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def draft_status_for_row(row: dict[str, str], draft: str) -> str:
    if not draft:
        return "empty"
    trial = dict(row)
    trial["index_summary"] = draft
    errors = [i.message for i in check_index_summary_row(trial) if i.level == "ERROR"]
    if errors:
        return "error: " + errors[0][:80]
    warns = [i.message for i in check_index_summary_row(trial) if i.level == "WARN"]
    if warns:
        return "warn: " + warns[0][:60]
    return "ok"


def build_review_rows(
    rows: list[dict[str, str]],
    *,
    prefill_final: bool = False,
) -> list[dict[str, str]]:
    entry_by_term = {e["term"]: e for e in load_glossary_entries()}
    seen_keys: set[str] = set()
    out: list[dict[str, str]] = []

    sorted_rows = sorted(
        rows,
        key=lambda r: (norm(r.get("category")), norm(r.get("term"))),
    )

    for row in sorted_rows:
        term = norm(row.get("term"))
        if not term:
            continue
        current = norm(row.get("index_summary"))
        draft = draft_glossary_index_summary(row, seen_keys=seen_keys)
        entry = entry_by_term.get(term, row)
        legacy = terms_index_snippet(entry) if not current else current
        faq = norm(row.get("faq_1_answer"))
        if len(faq) > 180:
            faq = faq[:177] + "…"
        status = draft_status_for_row(row, draft)
        final = current
        if prefill_final and not final and draft and (
            status == "ok" or status.startswith("warn")
        ):
            final = draft
        out.append(
            {
                "term": term,
                "category": norm(row.get("category")),
                "importance": norm(row.get("importance")),
                "index_summary_current": current,
                "index_summary_draft": draft,
                "index_summary_final": final,
                "legacy_snippet": legacy,
                "legal_basis": norm(row.get("legal_basis")),
                "exam_points": norm(row.get("exam_points")),
                "faq_1_excerpt": faq,
                "draft_len": str(len(draft)),
                "draft_status": status,
                "review_status": "confirmed" if current else ("draft" if final and not current else ""),
            }
        )
    return out


def export_review(path: Path, review_rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=REVIEW_COLUMNS, lineterminator="\n")
            w.writeheader()
            w.writerows(review_rows)
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "index_summary"
    ws.append(list(REVIEW_COLUMNS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in review_rows:
        ws.append([row.get(c, "") for c in REVIEW_COLUMNS])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 36
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 40
    ws.freeze_panes = "A2"
    wb.save(path)


def apply_drafts(
    rows: list[dict[str, str]],
    *,
    dry_run: bool,
    only_ok: bool,
) -> tuple[int, int, int]:
    seen_keys: set[str] = set()
    applied_pairs: list[tuple[str, str]] = []
    applied = skipped = failed = 0

    sorted_rows = sorted(
        rows,
        key=lambda r: (norm(r.get("category")), norm(r.get("term"))),
    )

    for row in sorted_rows:
        term = norm(row.get("term"))
        if not term:
            continue
        if norm(row.get("index_summary")):
            text = norm(row.get("index_summary"))
            seen_keys.add(normalize_key(text))
            if len(normalize_key(text)) >= 50:
                seen_keys.add(normalize_key(text)[:50])
            applied_pairs.append((term, text))
            skipped += 1
            continue

        draft = draft_glossary_index_summary(row, seen_keys=seen_keys)
        if not draft:
            failed += 1
            continue
        status = draft_status_for_row(row, draft)
        if only_ok and not status.startswith("ok"):
            failed += 1
            continue

        if cross_conflicts_with(term, draft, applied_pairs):
            draft2 = _dedupe_draft(term, draft, row, seen_keys)
            if draft2 and not cross_conflicts_with(term, draft2, applied_pairs):
                draft = draft2
            else:
                failed += 1
                continue

        if not dry_run:
            row["index_summary"] = draft
        applied_pairs.append((term, draft))
        applied += 1

    return applied, skipped, failed


def apply_all_ok_drafts(
    rows: list[dict[str, str]],
    *,
    dry_run: bool,
    include_warn: bool,
    force: bool,
) -> tuple[int, int, int]:
    """draft_status=ok（+任意で warn）の下書きを語間重複チェック付きで一括反映。"""
    seen_keys: set[str] = set()
    applied_pairs: list[tuple[str, str]] = []
    applied = skipped = failed = 0

    sorted_rows = sorted(
        rows,
        key=lambda r: (norm(r.get("category")), norm(r.get("term"))),
    )

    for row in sorted_rows:
        term = norm(row.get("term"))
        if not term:
            continue

        current = norm(row.get("index_summary"))
        if current and not force:
            seen_keys.add(normalize_key(current))
            if len(normalize_key(current)) >= 50:
                seen_keys.add(normalize_key(current)[:50])
            applied_pairs.append((term, current))
            skipped += 1
            continue

        draft = draft_glossary_index_summary(row, seen_keys=seen_keys)
        if not draft:
            failed += 1
            continue

        status = draft_status_for_row(row, draft)
        if status == "ok" or (include_warn and status.startswith("warn")):
            pass
        else:
            failed += 1
            continue

        if cross_conflicts_with(term, draft, applied_pairs):
            draft2 = _dedupe_draft(term, draft, row, seen_keys)
            if draft2 and not cross_conflicts_with(term, draft2, applied_pairs):
                draft = draft2
            else:
                failed += 1
                continue

        if not dry_run:
            row["index_summary"] = draft
        applied_pairs.append((term, draft))
        seen_keys.add(normalize_key(draft))
        if len(normalize_key(draft)) >= 50:
            seen_keys.add(normalize_key(draft)[:50])
        applied += 1

    return applied, skipped, failed


def _index_summary_rebuild_order(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """複合語・長い用語名を先に処理し、語間重複を起こしにくくする。"""
    return sorted(
        rows,
        key=lambda r: (
            norm(r.get("category")),
            -len(norm(r.get("term"))),
            norm(r.get("term")),
        ),
    )


def repair_cross_duplicate_summaries(
    rows: list[dict[str, str]],
    *,
    dry_run: bool,
    include_warn: bool,
) -> tuple[int, int]:
    """語間重複を解消するため全行を再生成して反映。"""
    seen_keys: set[str] = set()
    applied_pairs: list[tuple[str, str]] = []
    repaired = failed = 0

    sorted_rows = _index_summary_rebuild_order(rows)

    for row in sorted_rows:
        term = norm(row.get("term"))
        if not term:
            continue

        chosen = resolve_unique_index_summary(term, row, applied_pairs, seen_keys)
        if not chosen:
            chosen = _break_cross_similarity(term, row, applied_pairs)
        if not chosen:
            chosen = _minimal_exam_summary(term, row, applied_pairs)

        if not chosen:
            failed += 1
            continue

        status = draft_status_for_row(row, chosen)
        if status != "ok" and not (include_warn and status.startswith("warn")):
            alt = _break_cross_similarity(term, row, applied_pairs)
            if not alt:
                alt = _minimal_exam_summary(term, row, applied_pairs)
            if alt:
                chosen = alt
                status = draft_status_for_row(row, chosen)
            if status != "ok" and not (include_warn and status.startswith("warn")):
                failed += 1
                continue

        if not dry_run:
            row["index_summary"] = chosen
        _register_index_summary(term, chosen, applied_pairs, seen_keys)
        repaired += 1

    return repaired, failed


def _applied_pairs_without(
    rows: list[dict[str, str]],
    skip_term: str,
) -> tuple[list[tuple[str, str]], set[str]]:
    applied_pairs: list[tuple[str, str]] = []
    seen_keys: set[str] = set()
    for row in rows:
        term = norm(row.get("term"))
        summary = norm(row.get("index_summary"))
        if not term or not summary or term == skip_term:
            continue
        _register_index_summary(term, summary, applied_pairs, seen_keys)
    return applied_pairs, seen_keys


def polish_cross_duplicate_summaries(
    rows: list[dict[str, str]],
    *,
    dry_run: bool,
    include_warn: bool,
    max_passes: int = 5,
) -> int:
    """監査で検出した語間重複を対象語だけ差し替えて解消。"""
    by_term = {norm(r.get("term")): r for r in rows if norm(r.get("term"))}
    fixed_total = 0

    for _ in range(max_passes):
        errors = [
            i
            for i in audit_index_summary_cross_rows(rows)
            if i.level == "ERROR" and i.term
        ]
        if not errors:
            break

        terms_to_fix = list(dict.fromkeys(i.term for i in errors))
        pass_fixed = 0

        for term in terms_to_fix:
            row = by_term.get(term)
            if not row:
                continue

            applied_pairs, seen_keys = _applied_pairs_without(rows, term)
            partner = ""
            for issue in errors:
                if issue.term != term:
                    continue
                match = _CROSS_PARTNER_RE.search(issue.message)
                if match:
                    partner = match.group(1)
                    break

            chosen = ""
            if partner:
                for variant in range(20):
                    candidate = _build_substance_summary(
                        term, row, variant=variant, contrast_term=partner
                    )
                    chosen = _accept_candidate(
                        term, row, candidate, applied_pairs, build_mode=False
                    )
                    if chosen:
                        break
                    pushed = _push_below_cross_threshold(term, row, candidate, applied_pairs)
                    if pushed:
                        chosen = pushed
                        break

            if not chosen:
                chosen = _break_cross_similarity(term, row, applied_pairs)
            if not chosen:
                chosen = resolve_unique_index_summary(term, row, applied_pairs, seen_keys)
            if not chosen and partner:
                chosen = _minimal_exam_summary(
                    term, row, applied_pairs, contrast_term=partner
                )
            if not chosen:
                chosen = _minimal_exam_summary(term, row, applied_pairs)
            if not chosen or chosen == norm(row.get("index_summary")):
                continue

            status = draft_status_for_row(row, chosen)
            if status != "ok" and not (include_warn and status.startswith("warn")):
                continue

            if not dry_run:
                row["index_summary"] = chosen
            pass_fixed += 1

        fixed_total += pass_fixed
        if pass_fixed == 0:
            break

    return fixed_total


def fix_cross_error_terms(
    rows: list[dict[str, str]],
    *,
    dry_run: bool,
    include_warn: bool,
) -> int:
    """監査で語間重複 ERROR の用語だけを優先的に差し替える。"""
    from tools.index_summary_utils import _pad_to_min_length, split_semicolon  # noqa: E402

    by_term = {norm(r.get("term")): r for r in rows if norm(r.get("term"))}
    cross_terms = {
        i.term
        for i in audit_index_summary_cross_rows(rows)
        if i.level == "ERROR" and i.term
    }
    fixed = 0

    for term in sorted(cross_terms, key=lambda t: (-len(t), t)):
        row = by_term.get(term)
        if not row:
            continue

        applied_pairs, seen_keys = _applied_pairs_without(rows, term)
        chosen = resolve_unique_index_summary(term, row, applied_pairs, seen_keys)
        if not chosen:
            chosen = _break_cross_similarity(term, row, applied_pairs)
        if not chosen:
            chosen = _minimal_exam_summary(term, row, applied_pairs)

        if not chosen:
            category = norm(row.get("category"))
            legal = norm(row.get("legal_basis")).rstrip("。")
            eps = split_semicolon(norm(row.get("exam_points")))
            for variant in range(32):
                ep = (
                    eps[variant % len(eps)].rstrip("。")
                    if eps
                    else "定義と条件の言い換え"
                )
                seeds = [
                    (
                        f"試験では{term}を{category or '危険物'}の重要語として扱い、"
                        f"{ep}の理解が得点につながります"
                    ),
                    (
                        f"{term}は{legal or category or '危険物'}に関係し、"
                        f"試験では{ep}が問われます"
                    ),
                ]
                for seed in seeds:
                    padded = _pad_to_min_length(term, seed.rstrip("。"), row)
                    chosen = _accept_candidate(
                        term,
                        row,
                        padded,
                        applied_pairs,
                        build_mode=False,
                    )
                    if chosen:
                        break
                if chosen:
                    break

        if not chosen:
            continue

        status = draft_status_for_row(row, chosen)
        if status != "ok" and not (include_warn and status.startswith("warn")):
            continue

        if not dry_run:
            row["index_summary"] = chosen
        fixed += 1

    return fixed


def fill_remaining_summaries(
    rows: list[dict[str, str]],
    *,
    dry_run: bool,
    include_warn: bool,
) -> tuple[int, int]:
    """未記入行を seen_keys 付きで再生成し、語間重複を避けて反映。"""
    from tools.index_summary_utils import (  # noqa: E402
        _fallback_index_summary,
        _synthesize_opening,
        _trim_to_max,
        force_unique_index_summary,
    )

    seen_keys: set[str] = set()
    applied_pairs: list[tuple[str, str]] = []
    for row in rows:
        term = norm(row.get("term"))
        text = norm(row.get("index_summary"))
        if term and text:
            seen_keys.add(normalize_key(text))
            if len(normalize_key(text)) >= 50:
                seen_keys.add(normalize_key(text)[:50])
            applied_pairs.append((term, text))

    filled = 0
    failed = 0
    sorted_rows = sorted(
        rows,
        key=lambda r: (norm(r.get("category")), norm(r.get("term"))),
    )
    for row in sorted_rows:
        term = norm(row.get("term"))
        if not term or norm(row.get("index_summary")):
            continue

        chosen = ""
        resolved = resolve_unique_index_summary(term, row, applied_pairs, seen_keys)
        if resolved:
            chosen = resolved
        else:
            candidates: list[str] = []
            draft = draft_glossary_index_summary(row, seen_keys=seen_keys)
            if draft:
                candidates.append(draft)
            synth = _trim_to_max(_synthesize_opening(term, row).rstrip("。"))
            if synth:
                candidates.append(synth)
            fb = _fallback_index_summary(term, row)
            if fb:
                candidates.append(fb)
            forced = force_unique_index_summary(term, row)
            if forced:
                candidates.append(forced)

            for candidate in candidates:
                if not candidate:
                    continue
                status = draft_status_for_row(row, candidate)
                if status != "ok" and not (include_warn and status.startswith("warn")):
                    continue
                if cross_conflicts_with(term, candidate, applied_pairs):
                    alt = _dedupe_draft(term, candidate, row, seen_keys)
                    if alt and not cross_conflicts_with(term, alt, applied_pairs):
                        candidate = alt
                    else:
                        continue
                chosen = candidate
                break

        if not chosen:
            failed += 1
            continue
        if not dry_run:
            row["index_summary"] = chosen
        seen_keys.add(normalize_key(chosen))
        if len(normalize_key(chosen)) >= 50:
            seen_keys.add(normalize_key(chosen)[:50])
        applied_pairs.append((term, chosen))
        filled += 1

    return filled, failed


def apply_from_review(path: Path, rows: list[dict[str, str]], *, cross_safe: bool = True) -> int:
    if path.suffix.lower() == ".csv":
        review = list(csv.DictReader(path.open(encoding="utf-8-sig")))
    else:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        review = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            review.append(dict(zip(headers, [norm(v) if v is not None else "" for v in r])))

    by_term = {norm(r.get("term")): r for r in rows if norm(r.get("term"))}
    applied_pairs: list[tuple[str, str]] = [
        (norm(r.get("term")), norm(r.get("index_summary")))
        for r in rows
        if norm(r.get("term")) and norm(r.get("index_summary"))
    ]
    seen_keys: set[str] = set()
    for _, text in applied_pairs:
        seen_keys.add(normalize_key(text))
        if len(normalize_key(text)) >= 50:
            seen_keys.add(normalize_key(text)[:50])
    updated = 0
    review_sorted = sorted(
        review,
        key=lambda item: (norm(item.get("category")), norm(item.get("term"))),
    )
    for item in review_sorted:
        term = norm(item.get("term"))
        if not term or term not in by_term:
            continue
        final = norm(item.get("index_summary_final"))
        if not final:
            draft = norm(item.get("index_summary_draft"))
            status = norm(item.get("draft_status"))
            if draft and (status == "ok" or status.startswith("warn")):
                final = draft
        if not final:
            continue
        if by_term[term].get("index_summary") == final:
            continue
        if cross_safe and cross_conflicts_with(term, final, applied_pairs):
            regen = draft_glossary_index_summary(by_term[term], seen_keys=seen_keys)
            if regen and not cross_conflicts_with(term, regen, applied_pairs):
                final = regen
            else:
                continue
        by_term[term]["index_summary"] = final
        applied_pairs.append((term, final))
        seen_keys.add(normalize_key(final))
        if len(normalize_key(final)) >= 50:
            seen_keys.add(normalize_key(final)[:50])
        updated += 1
    return updated


def main() -> int:
    ap = argparse.ArgumentParser(description="用語一覧 index_summary の下書き・レビュー・反映")
    ap.add_argument("--export-review", metavar="PATH", help="レビュー用 Excel/CSV を出力")
    ap.add_argument("--draft", action="store_true", help="下書きを生成して CSV に反映")
    ap.add_argument("--apply", action="store_true", help="--draft 時に CSV を更新")
    ap.add_argument("--apply-from", metavar="PATH", help="レビュー済みファイルの index_summary_final を反映")
    ap.add_argument("--fill-remaining", action="store_true", help="未記入行だけ再生成して反映")
    ap.add_argument("--repair-cross", action="store_true", help="語間重複を解消するため全行を再生成")
    ap.add_argument("--apply-all-ok", action="store_true", help="ok 下書きを一括反映（--force で既存も上書き）")
    ap.add_argument("--force", action="store_true", help="既存 index_summary も下書きで置換")
    ap.add_argument("--prefill-final", action="store_true", help="--export-review 時に final 列へ下書きをコピー")
    ap.add_argument("--include-warn", action="store_true", help="draft_status=warn も --draft --apply 対象にする")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not CSV_PATH.is_file():
        print(f"missing: {CSV_PATH}", file=sys.stderr)
        return 1

    if not any([
        args.export_review,
        args.draft,
        args.apply_from,
        args.apply_all_ok,
        args.fill_remaining,
        args.repair_cross,
    ]):
        ap.error(
            "--export-review / --draft / --apply-from / --apply-all-ok / --fill-remaining"
            " / --repair-cross のいずれかを指定してください"
        )

    fieldnames, rows = load_rows()
    if "index_summary" not in fieldnames:
        print("index_summary 列がありません。先に列追加を実行してください。", file=sys.stderr)
        return 1

    if args.export_review:
        review_rows = build_review_rows(rows, prefill_final=args.prefill_final)
        out = Path(args.export_review).expanduser()
        export_review(out, review_rows)
        ok = sum(1 for r in review_rows if r["draft_status"] == "ok")
        warn = sum(1 for r in review_rows if r["draft_status"].startswith("warn"))
        err = sum(1 for r in review_rows if r["draft_status"].startswith("error") or r["draft_status"] == "empty")
        print(f"レビュー出力: {out}")
        print(f"下書き: ok {ok} / warn {warn} / error・empty {err}（目標 {INDEX_SUMMARY_MIN_LEN}〜{INDEX_SUMMARY_MAX_LEN} 字）")

    if args.draft:
        applied, skipped, failed = apply_drafts(
            rows,
            dry_run=not args.apply,
            only_ok=not args.include_warn,
        )
        mode = "dry-run" if not args.apply else "apply"
        print(f"下書き反映 ({mode}): 更新 {applied} / スキップ(既存) {skipped} / 見送り {failed}")

        if args.apply and not args.dry_run:
            save_rows(fieldnames, rows)
            print(f"Wrote {CSV_PATH}")

    if args.apply_all_ok:
        applied, skipped, failed = apply_all_ok_drafts(
            rows,
            dry_run=not args.apply,
            include_warn=args.include_warn,
            force=args.force,
        )
        mode = "dry-run" if not args.apply else "apply"
        print(
            f"一括反映 ({mode}): 更新 {applied} / スキップ {skipped} / 見送り {failed}"
            + ("（--force で既存上書き）" if args.force else "")
        )
        if args.apply and not args.dry_run:
            save_rows(fieldnames, rows)
            print(f"Wrote {CSV_PATH}")

    if args.fill_remaining:
        filled, failed = fill_remaining_summaries(
            rows,
            dry_run=not args.apply,
            include_warn=args.include_warn,
        )
        mode = "dry-run" if not args.apply else "apply"
        print(f"未記入行の反映 ({mode}): 更新 {filled} / 見送り {failed}")
        if args.apply and not args.dry_run:
            save_rows(fieldnames, rows)
            print(f"Wrote {CSV_PATH}")

    if args.repair_cross:
        repaired, failed = repair_cross_duplicate_summaries(
            rows,
            dry_run=not args.apply,
            include_warn=args.include_warn,
        )
        mode = "dry-run" if not args.apply else "apply"
        print(f"語間重複修復 ({mode}): 更新 {repaired} / 見送り {failed}")
        polished = polish_cross_duplicate_summaries(
            rows,
            dry_run=not args.apply,
            include_warn=args.include_warn,
        )
        if polished:
            print(f"語間重複ポリッシュ ({mode}): 更新 {polished}")
        targeted = fix_cross_error_terms(
            rows,
            dry_run=not args.apply,
            include_warn=args.include_warn,
        )
        if targeted:
            print(f"語間重複個別修復 ({mode}): 更新 {targeted}")
        if args.apply and not args.dry_run:
            save_rows(fieldnames, rows)
            print(f"Wrote {CSV_PATH}")

    if args.apply_from:
        path = Path(args.apply_from).expanduser()
        updated = apply_from_review(path, rows)
        if updated and not args.dry_run:
            save_rows(fieldnames, rows)
            print(f"index_summary_final 反映: {updated} 件 → {CSV_PATH}")
        else:
            print(f"index_summary_final 反映: {updated} 件（dry-run 相当: --apply-from のみで保存）")
            if updated:
                save_rows(fieldnames, rows)
                print(f"Wrote {CSV_PATH}")

    filled, total = index_summary_fill_stats(rows)
    print(f"index_summary 記入: {filled} / {total} 語")

    cross_errors = [i for i in audit_index_summary_cross_rows(rows) if i.level == "ERROR"]
    if cross_errors:
        print(f"語間重複 ERROR: {len(cross_errors)} 件（audit_glossary_index_summaries.py で詳細確認）", file=sys.stderr)

    if args.apply or args.apply_from or args.apply_all_ok or args.fill_remaining or args.repair_cross:
        print("Next: python3 tools/audit_glossary_index_summaries.py && python3 tools/build_glossary_pages.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
